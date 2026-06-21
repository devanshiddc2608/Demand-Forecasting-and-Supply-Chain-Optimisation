# src/build_excel_workbook.py
# ============================================================
# Builds the 5-sheet Excel scenario analysis workbook.
# Run: python src/build_excel_workbook.py
# ============================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.config import POWERBI_EXPORT_DIR, PROJECT_ROOT

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color="1F4E79")
THIN_BORDER = Border(*[Side(style="thin", color="B7B7B7")] * 4)

def style_header_row(ws, row_num, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

def autofit_columns(ws, df, start_col=1):
    for i, col in enumerate(df.columns):
        width = max(len(str(col)), df[col].astype(str).map(len).max()) + 3
        ws.column_dimensions[get_column_letter(start_col + i)].width = min(width, 35)

def write_df(ws, df, start_row=1, start_col=1, title=None):
    r = start_row
    if title:
        ws.cell(row=r, column=start_col, value=title).font = TITLE_FONT
        r += 2
    for j, col in enumerate(df.columns):
        ws.cell(row=r, column=start_col + j, value=col)
    style_header_row(ws, r, len(df.columns))
    for i, row in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row):
            ws.cell(row=r + i, column=start_col + j, value=val)
    autofit_columns(ws, df, start_col)
    return r + len(df) + 1


def build_workbook():
    inv = pd.read_csv(f"{POWERBI_EXPORT_DIR}/pb_inventory_policy.csv")
    scenario = pd.read_csv(f"{POWERBI_EXPORT_DIR}/pb_scenario_analysis.csv")

    wb = Workbook()

    # ---- SHEET 0: Executive Summary ----
    ws0 = wb.active
    ws0.title = "Executive Summary"
    ws0["B2"] = "Demand Forecasting & Supply Chain Optimisation"
    ws0["B2"].font = Font(bold=True, size=16, color="1F4E79")
    ws0["B3"] = "Rossmann Retail Network | Scenario Analysis Workbook"
    ws0["B5"] = "Total Current Safety Stock Cost (Annual)"
    ws0["D5"] = inv["Current_SS_Cost"].sum()
    ws0["B6"] = "Total Recommended Safety Stock Cost (Annual)"
    ws0["D6"] = inv["Annual_SS_Cost"].sum()
    ws0["B7"] = "Annual Cost Saving"
    ws0["D7"] = inv["Annual_Cost_Saving"].sum()
    ws0["B8"] = "Cost Reduction %"
    ws0["D8"] = inv["Annual_Cost_Saving"].sum() / inv["Current_SS_Cost"].sum()
    ws0["D8"].number_format = "0.0%"
    for r in [5,6,7]:
        ws0[f"D{r}"].number_format = "€#,##0"
    ws0.column_dimensions["B"].width = 40

    # ---- SHEET 1: Assumptions ----
    ws1 = wb.create_sheet("1. Assumptions")
    assumptions = pd.DataFrame({
        "Parameter": ["Lead Time (days)", "Lead Time Std Dev (days)",
                      "Holding Cost Rate (% of inventory value/year)",
                      "Ordering Cost (€ per order)", "Stockout Penalty Rate (% margin)",
                      "Service Level - A Class", "Service Level - B Class",
                      "Service Level - C Class"],
        "Value": [7, 2, "25%", 150, "30%", "99%", "95%", "90%"],
        "Realistic Range": ["3-14 days", "1-4 days", "18-30%", "€50-€300",
                             "15-40%", "97-99.9%", "92-98%", "85-95%"]
    })
    write_df(ws1, assumptions, title="Input Assumptions")

    # ---- SHEET 2: Current State ----
    ws2 = wb.create_sheet("2. Current State")
    current_state = (inv.groupby("ABC_Class")
                      .agg(Stores=("Store","count"),
                           Avg_Daily_Sales=("Avg_Daily_Sales","mean"),
                           Current_SS_Cost=("Current_SS_Cost","sum"),
                           Avg_Store_MAPE=("Store_MAPE","mean"))
                      .reset_index())
    write_df(ws2, current_state, title="Current Inventory Policy Performance by Class")

    # ---- SHEET 3: Forecast-Driven Policy ----
    ws3 = wb.create_sheet("3. Forecast-Driven Policy")
    policy_cols = ["Store","ABC_Class","XYZ_Class","Avg_Daily_Sales",
                    "Safety_Stock","Reorder_Point","EOQ_Days",
                    "Annual_SS_Cost","Annual_Cost_Saving"]
    write_df(ws3, inv[policy_cols].sort_values("Annual_Cost_Saving", ascending=False),
              title="Recommended Policy by Store (sorted by saving potential)")

    # ---- SHEET 4: Scenario Comparison ----
    ws4 = wb.create_sheet("4. Scenario Comparison")
    sc_summary = (scenario.groupby("Scenario")
                  .agg(Total_Safety_Stock=("Safety_Stock","sum"),
                       Total_Carrying_Cost=("Carrying_Cost","sum"),
                       Total_Stockout_Cost=("Stockout_Cost","sum"),
                       Total_Cost=("Total_Cost","sum"))
                  .reindex(["Base","Optimistic","Pessimistic"])
                  .reset_index())
    write_df(ws4, sc_summary, title="Cost Comparison Across Demand Scenarios")

    # ---- SHEET 5: Sensitivity Analysis ----
    ws5 = wb.create_sheet("5. Sensitivity Analysis")
    z_table = {0.90:1.282, 0.92:1.405, 0.95:1.645, 0.97:1.881,
               0.99:2.326, 0.995:2.576, 0.999:3.090}
    base_mu = inv["Avg_Daily_Sales"].mean()
    base_sigma = inv["Avg_Daily_Sales"].std()  # proxy
    sens_rows = []
    for sl, z in z_table.items():
        ss = z * (base_sigma * (7 ** 0.5))
        cost = ss * 0.25 * len(inv)  # scaled to full network
        sens_rows.append({"Service_Level": sl, "Z_Factor": z,
                           "Network_Safety_Stock_Cost": round(cost, 0)})
    sens_df = pd.DataFrame(sens_rows)
    write_df(ws5, sens_df, title="Cost Sensitivity to Service Level Target (90%–99.9%)")

    out_path = PROJECT_ROOT / "excel" / "scenario_analysis.xlsx"
    wb.save(out_path)
    print(f"Workbook saved: {out_path}")

if __name__ == "__main__":
    build_workbook()